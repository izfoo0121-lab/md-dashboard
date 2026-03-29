"""
save_history.py — Append current month's results to history.xlsx
Run after process_data.py each month end.
Creates history.xlsx if it doesn't exist.
"""
import json
import os
from pathlib import Path
from datetime import date, datetime
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR    = Path(__file__).parent
DATA_FILE   = BASE_DIR / "dashboard_data.json"
HISTORY_FILE = BASE_DIR / "history.xlsx"
TARGETS_FILE = BASE_DIR / "targets.json"

# ── Quarters ──────────────────────────────────────────────────────────────────
QUARTERS = {1:"Q1",2:"Q1",3:"Q1",4:"Q2",5:"Q2",6:"Q2",7:"Q3",8:"Q3",9:"Q3",10:"Q4",11:"Q4",12:"Q4"}

def log(msg): print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")

def get_quarter(month_label):
    """e.g. 'Mar 26' → Q1 2026"""
    try:
        dt = datetime.strptime(month_label, "%b %y")
        return f"{QUARTERS[dt.month]} {dt.year}"
    except:
        return "—"

def load_data():
    with open(DATA_FILE, encoding="utf-8") as f:
        return json.load(f)

def load_targets():
    try:
        with open(TARGETS_FILE, encoding="utf-8") as f:
            return json.load(f)
    except:
        return {}

def build_monthly_rows(data, targets):
    """Build rows for Monthly_Summary sheet."""
    month   = data.get("current_month", "")
    quarter = get_quarter(month)
    wd      = data.get("working_days", {})
    agents  = data.get("agents", {})
    rows    = []

    for agent, adata in agents.items():
        sp  = adata.get("sales_progression", {})
        bc  = adata.get("brand_commission", {})
        dc  = adata.get("debtor_cards", {})
        kpi = adata.get("kpi", {})
        ag_tgts = targets.get("agents", {}).get(agent, {})
        kpi_tgts = ag_tgts.get("kpi_targets", {})

        tiers = sp.get("tiers", {})
        t1    = tiers.get("normal_t1", {}) or {}
        ga_t  = tiers.get("ga", {}) or {}
        ma_t  = tiers.get("ma", {}) or {}

        debtors   = dc.get("debtors", [])
        new_acc   = sum(1 for d in debtors if d.get("is_new", False))
        vip_cnt   = sum(1 for d in debtors if d.get("vip", False))

        row = {
            "Year":            datetime.strptime(month, "%b %y").year if month else "",
            "Month":           month,
            "Quarter":         quarter,
            "Agent":           agent,
            "Working Days":    wd.get("total_working_days", 0),
            # Sales
            "Normal Target":   t1.get("target", 0) or 0,
            "Normal Actual":   sp.get("normal_ctn", 0) or 0,
            "Normal %":        round(t1.get("pct", 0) or 0, 1),
            "GA Target":       ga_t.get("target", 0) or 0,
            "GA Actual":       sp.get("ga_ctn", 0) or 0,
            "GA %":            round(ga_t.get("pct", 0) or 0, 1),
            "MA Target":       ma_t.get("target", 0) or 0,
            "MA Actual":       sp.get("ma_ctn", 0) or 0,
            "MA %":            round(ma_t.get("pct", 0) or 0, 1),
            "Total Canggih CTN": sp.get("total_canggih_ctn", 0) or 0,
            "8COM Paid CTN":   sp.get("eightcom_paid_ctn", 0) or 0,
            "Txn Count":       sp.get("txn_count", 0) or 0,
            "Avg Txn/Day":     sp.get("avg_txn_per_day", 0) or 0,
            # KPI metrics
            "Total Debtors":   dc.get("total_debtors", 0) or 0,
            "Active":          dc.get("active_count", 0) or 0,
            "Pending":         dc.get("pending_count", 0) or 0,
            "Reactivation":    dc.get("reactivation_count", 0) or 0,
            "Activation Rate %": dc.get("activation_rate", 0) or 0,
            "New Accounts":    new_acc,
            "VIP Count":       vip_cnt,
            "New SKU":         dc.get("total_new_sku", 0) or 0,
            # KPI targets
            "Target 开新户口":  kpi_tgts.get("new_accounts", 5) or 5,
            "Target VIP":      kpi_tgts.get("vip_count", 3) or 3,
            "Target 激活户口":  kpi_tgts.get("reactivation", 5) or 5,
            "Target 加SKU":     kpi_tgts.get("new_sku", 17) or 17,
            "Target 光顾率%":   kpi_tgts.get("activation_rate", 80) or 80,
            "Target Event":    kpi_tgts.get("event", 16) or 16,
            # Brand commission
            "Brands Earned":   sum(1 for b in bc.values() if b.get("status") == "both_hit"),
            "Total Comm RM":   round(sum(b.get("comm_earned", 0) for b in bc.values()), 2),
            # Brand commission detail per brand
            "iFACE Pen":   bc.get("iFACE",{}).get("new_penetrations", 0) or 0,
            "iFACE CTN":   bc.get("iFACE",{}).get("cur_ctn", 0) or 0,
            "iFACE Comm":  "✓" if bc.get("iFACE",{}).get("status")=="both_hit" else "✗",
            "SUKUN Pen":   bc.get("SUKUN",{}).get("new_penetrations", 0) or 0,
            "SUKUN CTN":   bc.get("SUKUN",{}).get("cur_ctn", 0) or 0,
            "SUKUN Comm":  "✓" if bc.get("SUKUN",{}).get("status")=="both_hit" else "✗",
            "EVO Pen":     bc.get("EVO",{}).get("new_penetrations", 0) or 0,
            "EVO CTN":     bc.get("EVO",{}).get("cur_ctn", 0) or 0,
            "EVO Comm":    "✓" if bc.get("EVO",{}).get("status")=="both_hit" else "✗",
            "BISON Pen":   bc.get("BISON",{}).get("new_penetrations", 0) or 0,
            "BISON CTN":   bc.get("BISON",{}).get("cur_ctn", 0) or 0,
            "BISON Comm":  "✓" if bc.get("BISON",{}).get("status")=="both_hit" else "✗",
            "TR20 Pen":    bc.get("TR20",{}).get("new_penetrations", 0) or 0,
            "TR20 CTN":    bc.get("TR20",{}).get("cur_ctn", 0) or 0,
            "TR20 Comm":   "✓" if bc.get("TR20",{}).get("status")=="both_hit" else "✗",
            "LAM+LWM Pen": bc.get("LAM+LWM",{}).get("new_penetrations", 0) or 0,
            "LAM+LWM CTN": bc.get("LAM+LWM",{}).get("cur_ctn", 0) or 0,
            "LAM+LWM Comm":"✓" if bc.get("LAM+LWM",{}).get("status")=="both_hit" else "✗",
            # KPI scores
            "KPI A+B+C Score": round(kpi.get("total_abc", 0) or 0, 2),
            "KPI Grand Total": round(kpi.get("grand_total", 0) or 0, 2),
            "KPI %":           round(kpi.get("total_pct", 0) or 0, 1),
            # Sections
            "KPI Sec A":       round((kpi.get("section_scores", {}).get("A", {}) or {}).get("score", 0), 2),
            "KPI Sec B":       round((kpi.get("section_scores", {}).get("B", {}) or {}).get("score", 0), 2),
            "KPI Sec C":       round((kpi.get("section_scores", {}).get("C", {}) or {}).get("score", 0), 2),
            "KPI Sec D":       kpi.get("section_d", {}).get("score") or "",
            "KPI Sec E":       kpi.get("section_e", {}).get("score") or "",
            "Saved At":        datetime.now().strftime("%Y-%m-%d %H:%M"),
        }
        rows.append(row)
    return rows

def build_sku_rows(data):
    """Build per-agent per-SKU per-month rows for SKU_History sheet."""
    month  = data.get("current_month", "")
    agents = data.get("agents", {})
    rows   = []
    for agent, adata in agents.items():
        sp = adata.get("sales_progression", {})
        sku_trend = sp.get("sku_trend", {})
        months = sp.get("month_labels", [])
        cur_m  = months[-1] if months else month
        mdata  = {sku: trend.get(cur_m, {}) for sku, trend in sku_trend.items()}
        for sku, d in mdata.items():
            rows.append({
                "Month":   month,
                "Agent":   agent,
                "SKU":     sku,
                "Debtors": d.get("debtors", 0),
                "CTN":     d.get("ctn", 0),
            })
    return rows


def build_team_rows(data, targets):
    """Build team-level monthly summary row."""
    month   = data.get("current_month", "")
    quarter = get_quarter(month)
    wd      = data.get("working_days", {})
    agents  = data.get("agents", {})
    team    = data.get("team_summary", {})
    gb      = data.get("group_brand_targets", {})

    # Aggregate agent stats
    total_normal_ctn = sum(a.get("sales_progression",{}).get("normal_ctn",0) or 0 for a in agents.values())
    total_canggih    = sum(a.get("sales_progression",{}).get("total_canggih_ctn",0) or 0 for a in agents.values())
    total_txns       = sum(a.get("sales_progression",{}).get("txn_count",0) or 0 for a in agents.values())
    total_active     = sum(a.get("debtor_cards",{}).get("active_count",0) or 0 for a in agents.values())
    total_debtors    = sum(a.get("debtor_cards",{}).get("total_debtors",0) or 0 for a in agents.values())
    total_new_acc    = sum(sum(1 for d in a.get("debtor_cards",{}).get("debtors",[]) if d.get("is_new")) for a in agents.values())
    total_new_sku    = sum(a.get("debtor_cards",{}).get("total_new_sku",0) or 0 for a in agents.values())

    row = {
        "Year":             datetime.strptime(month, "%b %y").year if month else "",
        "Month":            month,
        "Quarter":          quarter,
        "Working Days":     wd.get("total_working_days", 0),
        "Agents":           len(agents),
        "Team Normal CTN":  round(total_normal_ctn, 0),
        "Team Canggih CTN": round(total_canggih, 0),
        "Team Txn Count":   total_txns,
        "Team Active":      total_active,
        "Team Total Debtors": total_debtors,
        "Team Act Rate %":  round(total_active/total_debtors*100,1) if total_debtors else 0,
        "Team New Accts":   total_new_acc,
        "Team New SKU":     total_new_sku,
    }
    # Group brand actuals
    for brand, bdata in gb.items():
        row[f"GB {brand} CTN"]    = bdata.get("actual_ctn", 0) or 0
        row[f"GB {brand} Target"] = bdata.get("target_ctn", 0) or 0
        row[f"GB {brand} %"]      = bdata.get("pct", 0) or 0

    row["Saved At"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    return [row]



    """Build per-agent per-SKU per-month rows for SKU_History sheet."""
    month  = data.get("current_month", "")
    agents = data.get("agents", {})
    rows   = []
    for agent, adata in agents.items():
        sp = adata.get("sales_progression", {})
        sku_trend = sp.get("sku_trend", {})
        months = sp.get("month_labels", [])
        cur_m  = months[-1] if months else month
        mdata  = {sku: trend.get(cur_m, {}) for sku, trend in sku_trend.items()}
        for sku, d in mdata.items():
            rows.append({
                "Month":   month,
                "Agent":   agent,
                "SKU":     sku,
                "Debtors": d.get("debtors", 0),
                "CTN":     d.get("ctn", 0),
            })
    return rows

def style_header(ws, row=1, bg="1A1714", fg="FFFFFF"):
    """Style header row."""
    for cell in ws[row]:
        cell.font      = Font(name="Arial", bold=True, color=fg, size=9)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = Border(
            bottom=Side(style="thin", color="555555"),
            right=Side(style="thin",  color="555555"),
        )
    ws.row_dimensions[row].height = 30

def append_to_sheet(wb, sheet_name, new_rows, key_cols):
    """Append rows to sheet, avoiding duplicates by key columns."""
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        if new_rows:
            headers = list(new_rows[0].keys())
            ws.append(headers)
            style_header(ws)
            for col_idx, _ in enumerate(headers, 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 14
    else:
        ws = wb[sheet_name]

    # Get existing keys to avoid duplicates
    if ws.max_row > 1:
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
        key_indices = [headers.index(k)+1 for k in key_cols if k in headers]
        existing_keys = set()
        for r in range(2, ws.max_row+1):
            key = tuple(ws.cell(r, i).value for i in key_indices)
            existing_keys.add(key)
    else:
        headers = list(new_rows[0].keys()) if new_rows else []
        key_indices = [headers.index(k)+1 for k in key_cols if k in headers]
        existing_keys = set()

    added = 0
    for row in new_rows:
        vals   = list(row.values())
        h_list = list(row.keys())
        key    = tuple(row.get(k) for k in key_cols)
        if key in existing_keys:
            continue  # Skip duplicate
        ws.append(vals)
        # Colour alternate rows
        row_num = ws.max_row
        bg = "FAFAFA" if row_num % 2 == 0 else "FFFFFF"
        for c in range(1, len(vals)+1):
            ws.cell(row_num, c).fill = PatternFill("solid", fgColor=bg)
            ws.cell(row_num, c).font = Font(name="Arial", size=9)
            ws.cell(row_num, c).alignment = Alignment(horizontal="center", vertical="center")
        added += 1

    return added

def save_history():
    log("=" * 50)
    log("save_history.py — Monthly Snapshot")
    log("=" * 50)

    if not DATA_FILE.exists():
        log("❌ dashboard_data.json not found. Run process_data.py first.")
        return

    data    = load_data()
    targets = load_targets()
    month   = data.get("current_month", "unknown")
    log(f"Saving snapshot for: {month}")

    # Load or create workbook
    if HISTORY_FILE.exists():
        wb = load_workbook(HISTORY_FILE)
        log(f"  Loaded existing history.xlsx ({HISTORY_FILE.stat().st_size//1024} KB)")
    else:
        wb = Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        log("  Creating new history.xlsx")

    # ── Sheet 1: Monthly Summary ──────────────────────────────────────────
    monthly_rows = build_monthly_rows(data, targets)
    added = append_to_sheet(wb, "Monthly_Summary", monthly_rows, ["Month", "Agent"])
    log(f"  Monthly_Summary: +{added} rows (month={month})")

    # ── Sheet 2: SKU History ──────────────────────────────────────────────
    sku_rows = build_sku_rows(data)
    added = append_to_sheet(wb, "SKU_History", sku_rows, ["Month", "Agent", "SKU"])
    log(f"  SKU_History: +{added} rows")

    # ── Sheet 3: Team Summary ─────────────────────────────────────────────
    team_rows = build_team_rows(data, targets)
    added = append_to_sheet(wb, "Team_Summary", team_rows, ["Month"])
    log(f"  Team_Summary: +{added} rows")

    # ── Save ──────────────────────────────────────────────────────────────
    wb.save(HISTORY_FILE)
    size_kb = HISTORY_FILE.stat().st_size / 1024
    log(f"\n✅ history.xlsx saved — {size_kb:.0f} KB")
    log(f"   {len(monthly_rows)} agents · {month}")
    log("=" * 50)

if __name__ == "__main__":
    save_history()
