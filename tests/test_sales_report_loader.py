import inspect
import os
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl

import process_data


NEW_HEADERS = [
    "Tranx Mth",
    "Doc. No.",
    "Date",
    "Debtor Code",
    "Company Name",
    "Sales Agent",
    "Area Code",
    "Item Group",
    "Item Code",
    "Item Description",
    "UOM",
    "Smallest Qty",
    "Unit Price",
    "Discount",
    "Local SubTotal",
    "Rebate",
    "PAID ON",
    "Debtor Type",
    "UNIQ CODE",
    "RM / CTN",
    "RM / CTN (REBATE)",
    "Sales type",
    "Comm Rate",
    "QTY (CTN)",
    "QTY (MC)",
    "RM / MC",
    "> Shop Price Comm",
]


class SalesReportLoaderTests(unittest.TestCase):
    def _write_workbook(self, headers, row):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "MD"
        sheet.append(headers)
        sheet.append(row)
        handle = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        handle.close()
        workbook.save(handle.name)
        workbook.close()
        return handle.name

    def test_new_debtor_type_column_does_not_shift_sales_fields(self):
        row = [
            "Jul",
            "IV-001",
            "2026-07-02",
            "300-CONV",
            "Converter Shop",
            "CJ",
            "GRP 2A",
            "SUKUN",
            "SKNR",
            "Sukun Red",
            "CTN",
            12,
            40,
            0,
            480,
            0,
            "Jul 26",
            "Converter",
            "UNIQ-001",
            40,
            38.5,
            "Target",
            1.8,
            12,
            1,
            480,
            21.6,
        ]
        path = self._write_workbook(NEW_HEADERS, row)
        try:
            result = process_data.load_sales_report(path)
        finally:
            os.unlink(path)

        self.assertIn("debtor_type", result.columns)
        loaded = result.iloc[0]
        self.assertEqual(loaded["debtor_type"], "Converter")
        self.assertEqual(loaded["uniq_code"], "UNIQ-001")
        self.assertEqual(loaded["rm_ctn"], 40)
        self.assertEqual(loaded["rm_ctn_rebate"], 38.5)
        self.assertEqual(loaded["sales_type"], "Target")
        self.assertEqual(loaded["qty_ctn"], 12)
        self.assertIn("source_has_debtor_type", result.attrs)
        self.assertTrue(result.attrs["source_has_debtor_type"])

    def test_explicit_path_can_be_loaded_without_replacing_global_source(self):
        self.assertIn("path", inspect.signature(process_data.load_sales_report).parameters)
        row = [
            "Jul", "IV-002", "2026-07-03", "300-TEST", "Test Shop", "CJ",
            "GRP 2A", "EVO", "EVO", "EVO", "CTN", 1, 41, 0, 41, 0,
            "Jul 26", "Converter", "UNIQ-002", 41, 41, "Target", 1.8, 1,
            0, 0, 1.8,
        ]
        path = self._write_workbook(NEW_HEADERS, row)
        try:
            result = process_data.load_sales_report(path)
        finally:
            os.unlink(path)

        self.assertEqual(result.iloc[0]["debtor_code"], "300-TEST")

    def test_legacy_report_without_debtor_type_still_maps_by_header(self):
        headers = [header for header in NEW_HEADERS if header != "Debtor Type"]
        row = [
            "Jun", "IV-LEGACY", "2026-06-30", "300-OLD", "Legacy Shop", "CJ",
            "GRP 2A", "BISON", "BISON-R", "Bison", "CTN", 2, 45, 0, 90,
            0, "Jun 26", "UNIQ-OLD", 45, 44, "Target", 1.8, 2, 0, 0, 3.6,
        ]
        path = self._write_workbook(headers, row)
        try:
            result = process_data.load_sales_report(path)
        finally:
            os.unlink(path)

        loaded = result.iloc[0]
        self.assertEqual(loaded["debtor_type"], "")
        self.assertEqual(loaded["uniq_code"], "UNIQ-OLD")
        self.assertEqual(loaded["qty_ctn"], 2)
        self.assertIn("source_has_debtor_type", result.attrs)
        self.assertFalse(result.attrs["source_has_debtor_type"])
        quality = process_data.build_debtor_type_quality(result, None)
        self.assertFalse(quality["report_column_present"])

    def test_missing_required_header_raises_descriptive_schema_error(self):
        headers = [header for header in NEW_HEADERS if header != "QTY (CTN)"]
        row = ["x"] * len(headers)
        path = self._write_workbook(headers, row)
        try:
            with self.assertRaisesRegex(ValueError, "qty_ctn"):
                process_data.load_sales_report(path)
        finally:
            os.unlink(path)

    def test_source_path_resolver_prefers_environment_override(self):
        self.assertTrue(hasattr(process_data, "resolve_input_path"))
        with tempfile.TemporaryDirectory() as temp_dir:
            configured = Path(temp_dir) / "latest-sales.xlsx"
            with patch.dict(os.environ, {"MD_SALES_FILE": str(configured)}):
                resolved = process_data.resolve_input_path(
                    "MD_SALES_FILE", "MD Sales Report.xlsx"
                )

        self.assertEqual(resolved, configured)

    def test_source_file_metadata_records_path_size_and_timestamp(self):
        self.assertTrue(hasattr(process_data, "source_file_metadata"))
        handle = tempfile.NamedTemporaryFile(delete=False)
        try:
            handle.write(b"source-data")
            handle.close()
            metadata = process_data.source_file_metadata(handle.name)
        finally:
            os.unlink(handle.name)

        self.assertEqual(metadata["path"], str(Path(handle.name).resolve()))
        self.assertTrue(metadata["exists"])
        self.assertEqual(metadata["size_bytes"], 11)
        self.assertTrue(metadata["modified_at"])

    def test_debtor_loader_filters_all_group_master_to_group2a(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Code", "Company Name", "Debtor Type", "Area", "Agent", "Active"])
        sheet.append(["300-2A", "Group 2A Shop", "Converter", "GRP 2A", "CJ", "Checked"])
        sheet.append(["300-3", "Group 3 Shop", "Converter", "GRP 3", "JOSH", "Checked"])
        handle = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        handle.close()
        workbook.save(handle.name)
        workbook.close()
        try:
            debtors = process_data.load_debtors(handle.name)
        finally:
            os.unlink(handle.name)

        self.assertEqual(debtors["Code"].tolist(), ["300-2A"])

    def test_debtor_loader_rejects_master_without_area_scope_column(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Code", "Company Name", "Debtor Type", "Agent", "Active"])
        sheet.append(["300-UNKNOWN", "Unknown Scope", "Converter", "CJ", "Checked"])
        handle = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        handle.close()
        workbook.save(handle.name)
        workbook.close()
        try:
            with self.assertRaisesRegex(ValueError, "Area"):
                process_data.load_debtors(handle.name)
        finally:
            os.unlink(handle.name)

    def test_fast_loader_reuses_valid_source_fingerprinted_cache(self):
        signature = inspect.signature(process_data.load_sales_report).parameters
        self.assertIn("use_cache", signature)
        self.assertIn("cache_path", signature)
        row = [
            "Jul", "IV-CACHE", "2026-07-04", "300-CACHE", "Cache Shop", "CJ",
            "GRP 2A", "EVO", "EVO", "EVO", "CTN", 1, 41, 0, 41, 0,
            "Jul 26", "Converter", "UNIQ-CACHE", 41, 41, "Target", 1.8, 1,
            0, 0, 1.8,
        ]
        workbook_path = self._write_workbook(NEW_HEADERS, row)
        with tempfile.TemporaryDirectory() as temp_dir:
            cache_path = Path(temp_dir) / "sales.pkl"
            try:
                first = process_data.load_sales_report(
                    workbook_path, cache_path=cache_path
                )
                with patch.object(
                    process_data.pd,
                    "read_excel",
                    side_effect=AssertionError("Excel should not be read for a valid cache"),
                ):
                    second = process_data.load_sales_report(
                        workbook_path, use_cache=True, cache_path=cache_path
                    )
            finally:
                os.unlink(workbook_path)

        self.assertEqual(first.iloc[0]["debtor_code"], "300-CACHE")
        self.assertEqual(second.iloc[0]["debtor_type"], "Converter")


if __name__ == "__main__":
    unittest.main()
