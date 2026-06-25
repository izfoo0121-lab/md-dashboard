import unittest
from datetime import date
import os
import tempfile

import openpyxl
import pandas as pd

import process_data


class ProcessMonthOverrideTests(unittest.TestCase):
    def test_month_override_keeps_requested_month_and_lookback(self):
        cur_month, prev_months, explicit = process_data.resolve_runtime_months(
            today=date(2026, 6, 4),
            paid_on_vals=["May 26", "Jun 26"],
            argv=["--month", "May 26"],
        )

        self.assertTrue(explicit)
        self.assertEqual(cur_month, "May 26")
        self.assertEqual(prev_months, ["Apr 26", "Mar 26", "Feb 26"])

    def test_compact_month_override_is_normalized(self):
        cur_month, prev_months, explicit = process_data.resolve_runtime_months(
            today=date(2026, 6, 4),
            paid_on_vals=["May 26", "Jun 26"],
            argv=["--month", "may26"],
        )

        self.assertTrue(explicit)
        self.assertEqual(cur_month, "May 26")
        self.assertEqual(prev_months, ["Apr 26", "Mar 26", "Feb 26"])

    def test_without_override_auto_switches_to_latest_paid_month(self):
        cur_month, prev_months, explicit = process_data.resolve_runtime_months(
            today=date(2026, 7, 2),
            paid_on_vals=["Apr 26", "May 26"],
            argv=[],
        )

        self.assertFalse(explicit)
        self.assertEqual(cur_month, "May 26")
        self.assertEqual(prev_months, ["Apr 26", "Mar 26", "Feb 26"])

    def test_detect_sales_sheet_and_header_handles_archived_workbook(self):
        wb = openpyxl.Workbook()
        summary = wb.active
        summary.title = "Payable Sales"
        summary["A1"] = "Group Target"
        raw = wb.create_sheet("MD 31052026")
        headers = [
            "Tranx Mth", "Doc. No.", "Date", "Debtor Code", "Company Name",
            "Sales Agent", "Area Code", "Item Group", "Item Code", "Item Description",
            "UOM", "Smallest Qty", "Unit Price", "Discount", "Local SubTotal",
            "Rebate", "PAID ON", "DEBTOR TYPE", "RM / CTN", "RM / CTN (REBATE)",
            "Sales type", "Comm Rate", "QTY (CTN)",
        ]
        for col, value in enumerate(headers, start=1):
            raw.cell(row=2, column=col, value=value)

        handle = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        handle.close()
        try:
            wb.save(handle.name)
            sheet_name, header_row = process_data.detect_sales_sheet_and_header(handle.name)
        finally:
            os.unlink(handle.name)

        self.assertEqual(sheet_name, "MD 31052026")
        self.assertEqual(header_row, 1)

    def test_extract_payable_sales_summary_reads_group_and_brand_totals(self):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Payable Sales"
        sheet.cell(row=5, column=2, value="GRP 2")
        sheet.cell(row=5, column=3, value=64285)
        sheet.cell(row=5, column=4, value=7236)
        sheet.cell(row=5, column=5, value=56328)
        sheet.cell(row=5, column=6, value=9122)
        sheet.cell(row=5, column=7, value=63564)
        sheet.cell(row=5, column=9, value=47353)
        sheet.cell(row=5, column=10, value=12519)
        sheet.cell(row=5, column=11, value=3255)
        sheet.cell(row=13, column=2, value="GRP 2")
        sheet.cell(row=13, column=3, value=7984)
        sheet.cell(row=13, column=4, value=14337)
        sheet.cell(row=13, column=5, value=10426)

        handle = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        handle.close()
        try:
            wb.save(handle.name)
            summary = process_data.extract_payable_sales_summary(handle.name)
        finally:
            os.unlink(handle.name)

        self.assertEqual(summary["team_normal_ctn"], 47353)
        self.assertEqual(summary["team_ma_ctn"], 12519)
        self.assertEqual(summary["team_ga_ctn"], 3255)
        self.assertEqual(summary["cur_month_invoiced_paid"], 56328)
        self.assertEqual(summary["all_month_paid_ctn"], 63564)
        self.assertEqual(summary["brand_actuals"]["SUKUN"], 7984)
        self.assertEqual(summary["brand_actuals"]["CLASSMILD"], 14337)
        self.assertEqual(summary["brand_actuals"]["EVO"], 10426)

    def test_campaign_active_for_output_month_keeps_history_clean(self):
        june_active = {
            "active": True,
            "created_at": "2026-06-02T13:14:42+00:00",
            "deadline": "2026-06-30",
        }
        may_closed = {
            "active": False,
            "created_at": "2026-05-03T01:54:53+00:00",
            "deadline": "2026-05-31",
        }
        june_closed = {
            "active": False,
            "created_at": "2026-06-01T00:00:00+00:00",
            "deadline": "2026-06-30",
        }

        self.assertFalse(
            process_data.campaign_active_for_output_month(
                june_active, "May 26", today=date(2026, 6, 4)
            )
        )
        self.assertTrue(
            process_data.campaign_active_for_output_month(
                may_closed, "May 26", today=date(2026, 6, 4)
            )
        )
        self.assertFalse(
            process_data.campaign_active_for_output_month(
                june_closed, "Jun 26", today=date(2026, 6, 4)
            )
        )
        self.assertTrue(
            process_data.campaign_active_for_output_month(
                june_closed, "Jun 26", today=date(2026, 7, 1)
            )
        )

    def test_agent_scope_ignores_blank_non_scope_target_shells(self):
        targets = {
            "agents": {
                "BEN": {
                    "active": True,
                    "sales_progression": {"normal_t1": 900, "normal_t2": 1100},
                },
                "JW": {
                    "active": False,
                    "sales_progression": {"normal_t1": 700},
                },
                "CALSON": {
                    "active": True,
                    "sales_progression": {},
                    "brand_commission": {
                        "EVO": {"ctn_target": 0, "penetration_target": 0, "pen_auto": False},
                    },
                    "kpi_targets": {},
                    "campaign_targets": {},
                },
            }
        }
        scoped_sales = pd.DataFrame({"agent": ["BEN", "CALSON"]})

        all_agents, active_agents, inactive_agents = process_data.resolve_dashboard_agents(
            targets, scoped_sales
        )

        self.assertEqual(all_agents, ["BEN", "JW"])
        self.assertEqual(active_agents, ["BEN"])
        self.assertEqual(inactive_agents, ["JW"])


if __name__ == "__main__":
    unittest.main()
